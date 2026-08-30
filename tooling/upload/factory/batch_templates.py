"""Generate downloadable batch templates (CSV + XLSX)."""
from __future__ import annotations

import csv
import io
from typing import Tuple

from . import batch_schema as schema
from . import quota
from .audit import audit


def build_csv_template() -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=schema.COLUMNS, lineterminator="\n")
    writer.writeheader()
    for row in schema.EXAMPLE_ROWS:
        writer.writerow({c: row.get(c, "") for c in schema.COLUMNS})
    audit("template.downloaded", format="csv")
    return buf.getvalue().encode("utf-8-sig")


def build_xlsx_template() -> bytes:
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as e:
        raise RuntimeError("openpyxl is required to generate XLSX templates") from e

    wb = openpyxl.Workbook()

    # Instructions
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i["A1"] = "Aethelgard Batch Production"
    ws_i["A1"].font = Font(bold=True, size=16)
    for i, line in enumerate(schema.INSTRUCTIONS_TEXT.splitlines(), start=3):
        ws_i.cell(row=i, column=1, value=line)
    ws_i.column_dimensions["A"].width = 100

    snap = quota.snapshot()
    ws_i["A28"] = f"Current remaining quota (at template generation): {snap['remaining']} / {snap['limit']}"
    ws_i["A28"].font = Font(bold=True, color="B45309")

    # Batch Input
    ws = wb.create_sheet("Batch Input")
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    required_cols = {"listing_id", "concept"}
    for col_idx, name in enumerate(schema.COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
        cell.border = thin
        if name in required_cols:
            cell.comment = Comment("Required field", "Aethelgard")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(28, len(name) + 4))

    ws.freeze_panes = "A2"

    for r_idx, example in enumerate(schema.EXAMPLE_ROWS, start=2):
        for c_idx, name in enumerate(schema.COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=example.get(name, ""))
            cell.border = thin
            if name in required_cols:
                cell.fill = required_fill

    # Formula helper row
    help_row = 4
    ws.cell(row=help_row, column=1, value="")
    # Artwork total formula for example block (column artwork_count = 11)
    art_col = schema.COLUMNS.index("artwork_count") + 1
    ws.cell(row=6, column=1, value="Artwork total (example rows)")
    ws.cell(row=6, column=2, value=f"=SUM({get_column_letter(art_col)}2:{get_column_letter(art_col)}3)")
    ws.cell(row=7, column=1, value="Remaining quota warning")
    ws.cell(row=7, column=2, value=f"Keep total ≤ {snap['remaining']} today ({snap['limit']} daily max)")
    ws.cell(row=7, column=2).font = Font(color="B45309", bold=True)

    # Data validations
    dv_mode = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.ACQUISITION_MODES)) + '"',
        allow_blank=True,
    )
    dv_mode.error = "Choose a supported acquisition mode"
    dv_mode.errorTitle = "Invalid mode"
    ws.add_data_validation(dv_mode)
    mode_col = get_column_letter(schema.COLUMNS.index("acquisition_mode") + 1)
    dv_mode.add(f"{mode_col}2:{mode_col}500")

    dv_aspect = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.ASPECT_RATIOS)) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_aspect)
    aspect_col = get_column_letter(schema.COLUMNS.index("aspect_ratio") + 1)
    dv_aspect.add(f"{aspect_col}2:{aspect_col}500")

    dv_orient = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.ORIENTATIONS)) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_orient)
    orient_col = get_column_letter(schema.COLUMNS.index("orientation") + 1)
    dv_orient.add(f"{orient_col}2:{orient_col}500")

    dv_delivery = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.DELIVERY_TYPES)) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_delivery)
    del_col = get_column_letter(schema.COLUMNS.index("delivery_type") + 1)
    dv_delivery.add(f"{del_col}2:{del_col}500")

    dv_policy = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.SELECTION_POLICIES)) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_policy)
    pol_col = get_column_letter(schema.COLUMNS.index("selection_policy") + 1)
    dv_policy.add(f"{pol_col}2:{pol_col}500")

    dv_preset = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.STYLE_PRESETS)) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_preset)
    preset_col = get_column_letter(schema.COLUMNS.index("style_preset") + 1)
    dv_preset.add(f"{preset_col}2:{preset_col}500")

    dv_ptype = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(schema.PRODUCT_TYPES)) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_ptype)
    ptype_col = get_column_letter(schema.COLUMNS.index("product_type") + 1)
    dv_ptype.add(f"{ptype_col}2:{ptype_col}500")

    out = io.BytesIO()
    wb.save(out)
    audit("template.downloaded", format="xlsx")
    return out.getvalue()
