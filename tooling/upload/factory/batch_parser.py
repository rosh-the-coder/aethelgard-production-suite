"""CSV / XLSX batch parsing and validation."""
from __future__ import annotations

import csv
import io
import os
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from . import batch_schema as schema
from . import quota


def _norm_header(h: str) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def expand_artwork_units(rows: List[Dict[str, Any]]) -> int:
    total = 0
    for r in rows:
        try:
            n = int(float(r.get("artwork_count") or 1))
        except (TypeError, ValueError):
            n = 1
        total += max(1, n)
    return total


def parse_csv_bytes(data: bytes) -> List[Dict[str, str]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for raw in reader:
        row = {_norm_header(k): _as_str(v) for k, v in (raw or {}).items() if k}
        if not any(row.values()):
            continue
        rows.append(row)
    return rows


def parse_xlsx_bytes(data: bytes) -> List[Dict[str, str]]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is required for XLSX uploads") from e
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    # Prefer Batch Input sheet
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() in ("batch input", "batch_input", "input"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [_norm_header(h) for h in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for vals in rows_iter:
        row = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            val = vals[i] if i < len(vals) else None
            s = _as_str(val)
            if s:
                empty = False
            row[h] = s
        if empty:
            continue
        out.append(row)
    return out


def parse_upload(filename: str, data: bytes) -> List[Dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(data)
    if name.endswith(".csv"):
        return parse_csv_bytes(data)
    # sniff
    if data[:2] == b"PK":
        return parse_xlsx_bytes(data)
    return parse_csv_bytes(data)


def validate_rows(rows: List[Dict[str, str]]) -> Dict[str, Any]:
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    valid_rows: List[Dict[str, Any]] = []
    artwork_ids = set()
    listing_fields: Dict[str, Dict[str, str]] = {}

    listing_level_keys = (
        "listing_name",
        "product_type",
        "listing_title",
        "price_eur",
        "mockup_preset",
        "delivery_type",
        "tags",
    )

    for idx, raw in enumerate(rows, start=2):  # spreadsheet row (header=1)
        row = {c: _as_str(raw.get(c)) for c in schema.COLUMNS}
        # keep unknown passthrough notes
        for k, v in raw.items():
            if k not in row:
                row[k] = _as_str(v)

        row_errors = []
        row_warnings = []

        listing_id = row.get("listing_id") or ""
        if not listing_id:
            row_errors.append("missing listing_id")

        mode = (row.get("acquisition_mode") or "ai").lower()
        row["acquisition_mode"] = mode
        if mode not in schema.ACQUISITION_MODES:
            row_errors.append(f"unsupported acquisition_mode '{mode}'")

        if mode in ("ai", "graphic_poster") and not row.get("concept"):
            row_errors.append("missing concept for AI/poster mode")

        if mode == "public_domain" and not row.get("concept") and not row.get("artwork_title"):
            row_warnings.append("public_domain row has no concept/artwork_title search hint")

        preset = row.get("style_preset") or ""
        if preset and preset not in schema.STYLE_PRESETS:
            row_errors.append(f"unknown style_preset '{preset}'")

        orientation = (row.get("orientation") or "").lower()
        if orientation and orientation not in schema.ORIENTATIONS:
            row_errors.append(f"invalid orientation '{orientation}'")
        row["orientation"] = orientation

        aspect = row.get("aspect_ratio") or ""
        if aspect and aspect not in schema.ASPECT_RATIOS:
            row_errors.append(f"invalid aspect_ratio '{aspect}'")

        model = row.get("generation_model") or ""
        if model:
            resolved = schema.MODEL_ALIASES.get(model) or schema.MODEL_ALIASES.get(model.lower())
            if not resolved:
                row_errors.append(f"unknown model alias '{model}'")
            else:
                row["generation_model"] = resolved

        try:
            count = int(float(row.get("artwork_count") or 1))
            if count < 1:
                raise ValueError()
        except (TypeError, ValueError):
            row_errors.append("artwork_count must be a positive number")
            count = 1
        row["artwork_count"] = count

        product_type = (row.get("product_type") or "single").lower()
        row["product_type"] = product_type
        if product_type and product_type not in schema.PRODUCT_TYPES:
            row_errors.append(f"unknown product_type '{product_type}'")

        delivery = (row.get("delivery_type") or "drive_pdf").lower()
        row["delivery_type"] = delivery
        if delivery not in schema.DELIVERY_TYPES:
            row_errors.append(f"unsupported delivery_type '{delivery}'")

        policy = (row.get("selection_policy") or "").lower()
        if not policy:
            policy = "manual_review" if count > 1 else "first_success"
        if policy not in schema.SELECTION_POLICIES:
            row_errors.append(f"invalid selection_policy '{policy}'")
        row["selection_policy"] = policy

        price = row.get("price_eur") or ""
        if price:
            try:
                float(price)
            except ValueError:
                row_errors.append("non-numeric price_eur")

        art_id = row.get("artwork_id") or ""
        if art_id:
            if art_id in artwork_ids:
                row_errors.append(f"duplicate artwork_id '{art_id}'")
            artwork_ids.add(art_id)
        else:
            row_warnings.append("artwork_id missing — auto-assigned")
            row["artwork_id"] = f"auto_r{idx}"

        if listing_id:
            prev = listing_fields.get(listing_id)
            if prev is None:
                listing_fields[listing_id] = {k: row.get(k) or "" for k in listing_level_keys}
            else:
                for k in listing_level_keys:
                    a = prev.get(k) or ""
                    b = row.get(k) or ""
                    if a and b and a != b:
                        row_errors.append(
                            f"inconsistent listing field '{k}' for listing_id '{listing_id}' "
                            f"('{a}' vs '{b}')"
                        )

        row["_sheet_row"] = idx
        if row_errors:
            errors.append({"row": idx, "listing_id": listing_id, "errors": row_errors, "row_data": row})
        else:
            if row_warnings:
                warnings.append({"row": idx, "listing_id": listing_id, "warnings": row_warnings})
            valid_rows.append(row)

    artwork_total = expand_artwork_units(valid_rows) if not errors else expand_artwork_units(
        [e["row_data"] for e in errors] + valid_rows
    )
    # Prefer sum across all parsed for quota display
    all_normalized = []
    for e in errors:
        all_normalized.append(e["row_data"])
    all_normalized.extend(valid_rows)
    artwork_requested = expand_artwork_units(rows_as_units(rows))

    listings = group_by_listing(valid_rows)
    q = quota.can_accept(artwork_requested if not errors else expand_artwork_units(valid_rows))
    if artwork_requested > q["remaining"] and not any(
        "exceeds remaining daily quota" in str(e.get("errors")) for e in errors
    ):
        # blocking at batch level
        if artwork_requested > q["remaining"]:
            errors.append(
                {
                    "row": None,
                    "listing_id": None,
                    "errors": [
                        f"artwork total {artwork_requested} exceeds daily limit remaining "
                        f"({q['remaining']} of {q['limit']})"
                    ],
                }
            )

    blocking = len(errors) > 0
    return {
        "ok": not blocking,
        "blocking": blocking,
        "total_rows": len(rows),
        "valid_rows": len(valid_rows),
        "invalid_rows": len([e for e in errors if e.get("row") is not None]),
        "listings_detected": len(listings) if valid_rows else len({r.get("listing_id") for r in rows if r.get("listing_id")}),
        "artworks_requested": artwork_requested,
        "quota": q,
        "errors": errors,
        "warnings": warnings,
        "rows": valid_rows,
        "listings": listings,
        "grouping_preview": [
            {
                "listing_id": lid,
                "artwork_count": sum(int(r.get("artwork_count") or 1) for r in items),
                "modes": sorted({r.get("acquisition_mode") for r in items}),
                "product_type": (items[0].get("product_type") if items else "single"),
                "listing_name": (items[0].get("listing_name") if items else ""),
            }
            for lid, items in listings.items()
        ],
    }


def rows_as_units(raw_rows: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    """Normalize raw rows just enough to sum artwork_count."""
    out = []
    for raw in raw_rows:
        row = {c: _as_str(raw.get(c)) for c in schema.COLUMNS}
        try:
            row["artwork_count"] = max(1, int(float(row.get("artwork_count") or 1)))
        except (TypeError, ValueError):
            row["artwork_count"] = 1
        out.append(row)
    return out


def group_by_listing(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        groups[r.get("listing_id") or "unknown"].append(r)
    return dict(groups)


def validation_report_text(result: Dict[str, Any], filename: str = "") -> str:
    lines = [
        "Aethelgard Batch Validation Report",
        f"File: {filename}",
        f"OK: {result.get('ok')}",
        f"Rows: {result.get('total_rows')}",
        f"Valid: {result.get('valid_rows')}",
        f"Invalid: {result.get('invalid_rows')}",
        f"Listings: {result.get('listings_detected')}",
        f"Artworks requested: {result.get('artworks_requested')}",
        f"Quota remaining: {(result.get('quota') or {}).get('remaining')}",
        "",
        "ERRORS",
    ]
    for e in result.get("errors") or []:
        lines.append(f"- row {e.get('row')}: {', '.join(e.get('errors') or [])}")
    lines.append("")
    lines.append("WARNINGS")
    for w in result.get("warnings") or []:
        lines.append(f"- row {w.get('row')}: {', '.join(w.get('warnings') or [])}")
    return "\n".join(lines)
